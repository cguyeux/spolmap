import copy
from tkinter import *

import itertools
import more_itertools
import numpy as np
from openpyxl import load_workbook
from pickle import load
from sklearn.manifold import TSNE

with open('data/sit_to_lineage.pkl', 'rb') as f:
    my_sit = load(f)

with open('data/sit_to_sra.pkl', 'rb') as f:
    my_sra = load(f)


class Spolmap:
    def __init__(self):
        print("Reading spoligotypes")
        self._spoligo = self._get_spol()
        self._known_sit = self._get_lineage()
        print("Computing gaps")
        self._spol_gaps = self._get_gaps()

    def _get_spol(self):
        wb = load_workbook(filename='docs/SITVIT23882_PHELAN_SNPBASEDLIN_SORTED.xlsx',
                           read_only=True)
        ws = wb['new sheet']
        spoligos = {}
        for row in ws.iter_rows(min_row=2):
            if row[1].value != None:
                spol, sit = row[3].value.replace('n', '\u25A0').replace('o', '\u25A1'), row[9].value
                # if 'RR' in srr and sit not in spoligos:
                spoligos[sit] = spol
        '''for row in range(1, ws.nrows):
            srr, spol, sit = ws.cell_value(row, 1), ws.cell_value(row, 3).replace('n', '\u25A0').replace('o', '\u25A1'), ws.cell_value(row, 9)
            #if 'RR' in srr and sit not in spoligos:
            spoligos[sit] = spol'''
        return spoligos

    def _get_lineage(self):
        with open('data/lineage_to_sit.pkl', 'rb') as f:
            dico = load(f)
        return dico

    def _get_gaps(self):
        def get_gaps(dd):
            ee = [list(u) for u in
                  more_itertools.consecutive_groups([-1] + [k for k in range(len(dd)) if dd[k] == '■'] + [43])]
            return [(ee[u][-1] + 2, ee[u + 1][0]) for u in range(len(ee) - 1)]
        gaps = sorted(set(list(itertools.chain.from_iterable([get_gaps(self._spoligo[k]) for k in self._spoligo]))))
        spol_gaps = {}
        for k in self._spoligo:
            spol_gaps[k] = [int(u in get_gaps(self._spoligo[k])) for u in gaps]
        return spol_gaps

    def _getPosition(self, event):
        x,y = event.x, event.y
        if self._starting_window:
            try:
                self._plot()
                self._c.create_rectangle(self._start_window[0], self._start_window[1], x, y)
            except:
                pass
        for k in self._spoligo:
            if k in self._NODES:
                if abs(self._NODES[k][0]-x)<=4 and abs(self._NODES[k][1]-y)<=4:
                    txt = f"SIT:{k}"+'\n'
                    txt += self._spoligo[k]+"\n"
                    """for l in self._known_sit:
                        if k in self._known_sit[l]:
                            txt += l+'\n'"""
                    if k in my_sit:
                        txt += f"lineage:{', '.join(my_sit[k])}"
                    txt += '\n'
                    for l in my_sra[str(k)][:10]:
                        txt += l+'\n'
                    self.position.set(txt)
                    break

    def _startWindow(self, event):
        self._starting_window = True
        self._start_window = (event.x, event.y)

    def _stopWindow(self, event):
        self._stop_window = (event.x, event.y)
        self._starting_window = False
        nodes = [k for k in self._NODES
                 if (self._start_window[0] <= self._NODES[k][0] <= self._stop_window[0]
                 or self._start_window[0] >= self._NODES[k][0] >= self._stop_window[0])
                 and (self._start_window[1] <= self._NODES[k][1] <= self._stop_window[1]
                 or self._start_window[1] >= self._NODES[k][1] >= self._stop_window[1])]

        minimum_x = min([self._NODES_ORIGIN[k][0] for k in nodes])
        minimum_y = min([self._NODES_ORIGIN[k][1] for k in nodes])
        maximum_x = max([self._NODES_ORIGIN[k][0] for k in nodes])
        maximum_y = max([self._NODES_ORIGIN[k][1] for k in nodes])
        self._NODES = {}
        for k in nodes:
            self._NODES[k] = [int(((self._NODES_ORIGIN[k][0] + abs(minimum_x))*(int(self._c.cget('width'))-20))/(maximum_x-minimum_x))+10,
                              int(((self._NODES_ORIGIN[k][1] + abs(minimum_y))*(int(self._c.cget('height'))-self._root.winfo_rooty()-20))/(maximum_y-minimum_y))+10]
        self._plot()

    def _stopWindow2(self, event):
        self._stop_window = (event.x, event.y)
        self._starting_window = False
        self._plot()
        nodes = [k for k in self._NODES
                 if (self._start_window[0] <= self._NODES[k][0] <= self._stop_window[0]
                 or self._start_window[0] >= self._NODES[k][0] >= self._stop_window[0])
                 and (self._start_window[1] <= self._NODES[k][1] <= self._stop_window[1]
                 or self._start_window[1] >= self._NODES[k][1] >= self._stop_window[1])]
        print(nodes)

    def _color(self, kk):
        k = str(kk)
        if k in my_sit:
            # L1
            if [l for l in my_sit[k] if l.startswith('1')] != []:
                return 'pink'
            # L2
            elif [l for l in my_sit[k] if l.startswith('2')] != []:
                return 'blue'
            # L3
            elif  [l for l in my_sit[k] if l.startswith('3')] != []:
                return 'violet'
            # L4
            elif [l for l in my_sit[k] if l.startswith('4')] != []:
                return 'red'
            # L5
            elif [l for l in my_sit[k] if l.startswith('5')] != []:
                return 'brown'
            # L6
            elif  [l for l in my_sit[k] if l.startswith('6')] != []:
                return 'green'
            # L7
            elif  [l for l in my_sit[k] if l.startswith('7')] != []:
                return 'yellow'
            # orygis
            elif  [l for l in my_sit[k] if 'orygis' in l] != []:
                return 'darkgreen'
            # BCG
            elif k in ['482', '3025', '1037']:
                return 'lime'
            # caprae
            elif [l for l in my_sit[k] if 'caprae' in l] != []:
                return 'mediumseagreen'
            # microti
            elif [l for l in my_sit[k] if 'microti' in l] != []:
                return 'azure'
            # pinnipedii
            elif [l for l in my_sit[k] if 'pinnipedi' in l] != []:
                return 'lightcyan'
            # mungi
            elif [l for l in my_sit[k] if 'mungi' in l] != []:
                return 'teal'
            # capensis
            elif [l for l in my_sit[k] if 'capensis' in l] != []:
                return 'turquoise'
            # suricatta
            elif [l for l in my_sit[k] if 'surica' in l] != []:
                return 'darkturquoise'
            # canettii
            elif [l for l in my_sit[k] if 'canetti' in l] != []:
                return 'black'
            else:
                return 'grey'
        else:
            return 'grey'

    def _plot(self):
        self._c.delete("all")
        for item, k in enumerate(self._NODES):
            if self._color(k) == 'grey':
                self._c.create_oval(self._NODES[k][0]-2,
                          self._NODES[k][1]-2,
                          self._NODES[k][0]+2,
                          self._NODES[k][1]+2,
                          fill=self._color(k))
        for item, k in enumerate(self._NODES):
            if self._color(k) != 'grey':
                self._c.create_oval(self._NODES[k][0]-4,
                          self._NODES[k][1]-4,
                          self._NODES[k][0]+5,
                          self._NODES[k][1]+4,
                          fill=self._color(k))

    def _reinitiate(self):
        self._NODES = copy.deepcopy(self._NODES_ORIGIN)
        minimum_x = min([self._NODES[k][0] for k in self._NODES])
        minimum_y = min([self._NODES[k][1] for k in self._NODES])
        maximum_x = max([self._NODES[k][0] for k in self._NODES])
        maximum_y = max([self._NODES[k][1] for k in self._NODES])

        for k in self._NODES:
            self._NODES[k][0] = int(((self._NODES[k][0] + abs(minimum_x))*(int(self._c.cget('width'))-20))/(maximum_x-minimum_x))+10
            self._NODES[k][1] = int(((self._NODES[k][1] + abs(minimum_y))*(int(self._c.cget('height'))-self._root.winfo_rooty()-20))/(maximum_y-minimum_y))+10

        self._plot()

    def _interface(self):
        # Construction de la fenêtre principale «root»
        self._root = Tk()
        self._root.title('Spolmap')

        sw = self._root.winfo_screenwidth()
        sh = self._root.winfo_screenheight()
        self._root.geometry("%dx%d+0+0" % (sw, sh))
        self._root.resizable(False, False)
        self._root.update()

        print("Dimensionality reduction")
        X = np.array([self._spol_gaps[k] for k in self._spol_gaps])
        self._X = TSNE(learning_rate=200).fit_transform(X)
        self._NODES_ORIGIN = {}
        for k in range(len(self._spoligo)):
            self._NODES_ORIGIN[list(self._spoligo)[k]] = list(self._X[k,:])
        print("Reduction done")

        self._c = Canvas(self._root,
                   height=sh,
                   width=int(sw*0.7),
                   background='#ffffff')

        self._c.grid(row=0, rowspan=50, column=0)

        self._NODES = copy.deepcopy(self._NODES_ORIGIN)
        minimum_x = min([self._NODES[k][0] for k in self._NODES])
        minimum_y = min([self._NODES[k][1] for k in self._NODES])
        maximum_x = max([self._NODES[k][0] for k in self._NODES])
        maximum_y = max([self._NODES[k][1] for k in self._NODES])

        for k in self._NODES:
            self._NODES[k][0] = int(((self._NODES[k][0] + abs(minimum_x))*(int(self._c.cget('width'))-20))/(maximum_x-minimum_x))+10
            self._NODES[k][1] = int(((self._NODES[k][1] + abs(minimum_y))*(int(self._c.cget('height'))-self._root.winfo_rooty()-20))/(maximum_y-minimum_y))+10

        cpt = 0
        b = Button(self._root, text="Reinitiate view", command=self._reinitiate)
        b.grid(row=cpt, column=1,sticky=W+N)
        cpt += 1

        self.frame2 = LabelFrame(self._root,
                                 height=sh,
                                 width = int(sw*0.29))
        self.frame2.grid(row=cpt, column=1,sticky=W+N)

        self.position = StringVar()
        self.position.set("Coucou")
        textePosition = Label(self.frame2,
                              textvariable=self.position,
                              justify=LEFT)
        textePosition.grid(row = cpt, column = 1, sticky=N+W)
        self._starting_window = True
        self._plot()

        self._c.bind("<Motion>", self._getPosition)
        self._c.bind("<Button-1>", self._startWindow)
        self._c.bind("<ButtonRelease-1>", self._stopWindow)
        self._c.bind("<Button-3>", self._startWindow)
        self._c.bind("<ButtonRelease-3>", self._stopWindow2)
        # Lancement de la «boucle principale»
        self._root.mainloop()

    def run(self):
        self._interface()

if __name__ == '__main__':
    Spolmap().run()